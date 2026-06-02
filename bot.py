"""
飞书 Bot — 手机遥控电脑
接收飞书消息 → 执行任务 → 回复结果
"""

import json
import os
import re
import subprocess
import time
import hashlib
import requests
from pathlib import Path
from datetime import datetime
from flask import Flask, request, jsonify
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import shutil

# ── 加载配置 ──────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
with open(BASE_DIR / "config.json", encoding="utf-8") as f:
    CONFIG = json.load(f)

APP_ID = CONFIG["app_id"]
APP_SECRET = CONFIG["app_secret"]
PORT = CONFIG["port"]

# ── 飞书 API 辅助 ──────────────────────────────────────────────

def get_tenant_access_token():
    """获取 tenant_access_token，带缓存"""
    cache_file = BASE_DIR / ".token_cache"
    if cache_file.exists():
        data = json.loads(cache_file.read_text())
        if data.get("expire", 0) > time.time():
            return data["token"]

    resp = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": APP_ID, "app_secret": APP_SECRET},
        timeout=10,
    ).json()

    token = resp.get("tenant_access_token", "")
    cache_file.write_text(json.dumps({
        "token": token,
        "expire": time.time() + 7000,
    }))
    return token


def send_message(receive_id, text):
    """回复用户消息"""
    token = get_tenant_access_token()
    body = {
        "receive_id": receive_id,
        "msg_type": "text",
        "content": json.dumps({"text": text}),
    }
    resp = requests.post(
        "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=open_id",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=body,
        timeout=10,
    )
    return resp.json()


def send_card(receive_id, title, content):
    """发送富文本卡片消息"""
    token = get_tenant_access_token()
    card = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": title},
            "template": "blue",
        },
        "elements": [
            {"tag": "markdown", "content": content},
        ],
    }
    body = {
        "receive_id": receive_id,
        "msg_type": "interactive",
        "content": json.dumps(card),
    }
    resp = requests.post(
        "https://open.feishu.cn/open-apis/im/v1/messages?receive_id_type=open_id",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=body,
        timeout=10,
    )
    return resp.json()


# ── Coros 桥接 ───────────────────────────────────────────────────

_coros_activity_cache = []  # 缓存最近一次活动列表，用于按序号查详情/下载


def run_coros_cmd(*args):
    """调用 Node.js 桥接脚本，返回 {ok, ...} 或 {ok:false, error}"""
    script = BASE_DIR / "coros_api.mjs"
    try:
        result = subprocess.run(
            ["node", str(script)] + list(args),
            capture_output=True, text=True, timeout=30,
            cwd=str(BASE_DIR),
        )
        data = json.loads(result.stdout)
        if not data.get("ok"):
            stderr = result.stderr.strip()
            return {"ok": False, "error": data.get("error", stderr or "未知错误")}
        return data
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": "Coros 请求超时"}
    except json.JSONDecodeError:
        return {"ok": False, "error": f"桥接脚本异常: {result.stderr.strip()}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def format_duration(sec):
    """秒数转可读时间"""
    if not sec or sec <= 0:
        return "—"
    h, m = divmod(int(sec), 3600)
    m, s = divmod(m, 60)
    if h:
        return f"{h}h{m:02d}m{s:02d}s"
    return f"{m}m{s:02d}s"


def format_distance(meters):
    """米转可读距离"""
    if not meters:
        return "—"
    if meters >= 1000:
        return f"{meters / 1000:.2f} km"
    return f"{meters:.0f} m"


SPORT_NAMES = {
    100: "🏃 跑步", 101: "🏃 室内跑", 102: "🏔 越野跑", 103: "🏟 跑道跑",
    104: "🥾 徒步", 900: "🚶 步行",
    200: "🚴 公路骑行", 299: "🚴 骑行", 201: "🚴 室内骑行", 202: "🚵 山地骑行",
    203: "🚴 骑行", 204: "🚴 骑行", 205: "🚴 骑行",
    300: "🏊 泳池游泳", 301: "🏊 公开水域",
    400: "🏋 有氧", 401: "🏋 GPS有氧", 402: "💪 力量训练",
}


def sport_name(st):
    return SPORT_NAMES.get(st, f"运动({st})")


# ── 任务执行器 ──────────────────────────────────────────────────

WORK_DIR = Path("D:/feishu-tasks")
WORK_DIR.mkdir(parents=True, exist_ok=True)

DAILY_REPORT_DIR = Path("E:/于跃龙/每日工作计划/2026")
DAILY_REPORT_TEMPLATE = DAILY_REPORT_DIR / "2026.6.1宏名公司员工每日工作总结计划表.xlsx"


def cmd_help():
    return (
        "📋 **可用指令**\n"
        "/帮助 — 显示此帮助\n"
        "/计划 [内容] — 生成工作计划文件\n"
        "/下载 [URL] [文件名] — 下载文件\n"
        "/处理 [文件名] [操作] — 处理文件（暂支持：统计行数、提取文字）\n"
        "/文件 — 列出已下载/生成的文件\n"
        "/状态 — 查看电脑运行状态\n"
        "/删除 [文件名] — 删除指定文件\n"
        "/日报 [任务 - 状态; ...] — 生成每日工作日报表格（WPS）\n"
        "/coros帮助 — ⌚ Coros 运动数据功能"
    )


def cmd_plan(content):
    """生成工作计划文本文件"""
    filename = f"计划_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    filepath = WORK_DIR / filename
    plan_text = (
        f"工作计划\n"
        f"创建时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"{'─' * 30}\n\n"
        f"{content}\n\n"
        f"{'─' * 30}\n"
        f"※ 此计划由手机远程创建"
    )
    filepath.write_text(plan_text, encoding="utf-8")
    return f"✅ 计划已生成：`{filename}`"


def cmd_daily_report(args_str):
    """生成每日工作日报表格 — 复制模板并填入内容"""
    # 解析：任务1 - 状态1; 任务2 - 状态2（支持中英文分号）
    tasks = []
    args_str = args_str.replace("；", ";")
    for part in args_str.split(";"):
        part = part.strip()
        if not part:
            continue
        if " - " in part:
            content, status = part.split(" - ", 1)
            tasks.append((content.strip(), status.strip()))
        else:
            tasks.append((part.strip(), "待完成"))

    if not tasks:
        return "❌ 请至少输入一条任务\n格式：/日报 任务内容 - 完成状态"

    tasks = tasks[:3]  # 最多3条（C5-C7）

    today = datetime.now()
    date_str = f"{today.year}.{today.month}.{today.day}"
    filename = f"{date_str}宏名公司员工每日工作总结计划表.xlsx"
    filepath = DAILY_REPORT_DIR / filename

    # 复制模板
    if not DAILY_REPORT_TEMPLATE.exists():
        return f"❌ 模板文件不存在：{DAILY_REPORT_TEMPLATE}"
    shutil.copy(str(DAILY_REPORT_TEMPLATE), str(filepath))

    # 编辑
    wb = openpyxl.load_workbook(str(filepath))
    ws = wb.active

    # 更新 A2 日期
    ws["A2"] = f"日期：{date_str}  岗位：  姓名：于跃龙"

    # 填入任务 C5-C7 / D5-D7
    for i, (content, status) in enumerate(tasks):
        row = 5 + i
        ws[f"C{row}"] = content
        ws[f"D{row}"] = status

    wb.save(str(filepath))

    return (
        f"✅ 日报已生成：{filename}\n"
        f"📋 共 {len(tasks)} 条任务\n"
        f"📁 E:\\于跃龙\\每日工作计划\\2026\\"
    )


def cmd_download(url, filename=None):
    """下载文件到工作目录"""
    if not filename:
        filename = url.rstrip("/").split("/")[-1] or "download"
    filepath = WORK_DIR / filename
    try:
        resp = requests.get(url, stream=True, timeout=60, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        resp.raise_for_status()
        total = int(resp.headers.get("content-length", 0))
        with open(filepath, "wb") as f:
            downloaded = 0
            for chunk in resp.iter_content(8192):
                f.write(chunk)
                downloaded += len(chunk)
        size_mb = filepath.stat().st_size / (1024 * 1024)
        return f"✅ 下载完成：`{filename}` ({size_mb:.1f} MB)"
    except Exception as e:
        return f"❌ 下载失败：{e}"


def cmd_process(filename, operation="统计"):
    """处理文件"""
    filepath = WORK_DIR / filename
    if not filepath.exists():
        return f"❌ 文件不存在：`{filename}`"

    try:
        text = filepath.read_text(encoding="utf-8", errors="ignore")

        if operation in ("统计", "stats"):
            lines = text.count("\n") + 1
            chars = len(text)
            words = len(text.split())
            return (
                f"📊 **{filename}** 统计结果\n"
                f"行数：{lines}\n"
                f"字符数：{chars}\n"
                f"单词数：{words}\n"
                f"大小：{filepath.stat().st_size / 1024:.1f} KB"
            )
        elif operation in ("提取", "extract"):
            preview = text[:2000]
            return f"📄 **{filename}** 内容预览：\n```\n{preview}\n```"
        else:
            return f"❌ 不支持的操作：{operation}，支持：统计、提取"
    except Exception as e:
        return f"❌ 处理失败：{e}"


def cmd_files():
    """列出工作目录文件"""
    files = sorted(WORK_DIR.iterdir(), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return "📁 暂无文件"

    lines = ["📁 **文件列表**"]
    for f in files[:20]:
        mtime = datetime.fromtimestamp(f.stat().st_mtime).strftime("%m/%d %H:%M")
        size_kb = f.stat().st_size / 1024
        lines.append(f"• {f.name} — {size_kb:.1f}KB — {mtime}")
    return "\n".join(lines)


def cmd_status():
    """系统状态"""
    import platform
    return (
        f"🖥 **电脑状态**\n"
        f"系统：{platform.system()} {platform.release()}\n"
        f"时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"工作文件数：{len(list(WORK_DIR.iterdir()))}"
    )


def cmd_delete(filename):
    filepath = WORK_DIR / filename
    if filepath.exists():
        filepath.unlink()
        return f"✅ 已删除：`{filename}`"
    return f"❌ 文件不存在：`{filename}`"


# ── Coros 指令 ───────────────────────────────────────────────────

def cmd_coros_login(email, password):
    """配置 Coros 账号，保存凭据并登录"""
    creds_file = BASE_DIR / "coros_creds.json"
    if email and password:
        creds_file.write_text(json.dumps({"email": email, "password": password}), encoding="utf-8")
    elif not creds_file.exists():
        return "❌ 请提供账号密码：/coros登录 <邮箱> <密码>"

    result = run_coros_cmd("login", email or "", password or "")
    if result["ok"]:
        user = result.get("user", {})
        nickname = user.get("nickname", "?")
        return f"✅ Coros 登录成功！\n用户：{nickname}"
    return f"❌ Coros 登录失败：{result['error']}"


def cmd_coros_activities(page=1):
    """获取活动列表"""
    global _coros_activity_cache
    result = run_coros_cmd("activities", "--size", "10", "--page", str(page))
    if not result["ok"]:
        return f"❌ 获取活动失败：{result['error']}"

    data_list = result.get("dataList", [])
    total = result.get("count", 0)
    _coros_activity_cache = data_list

    if not data_list:
        return "📭 暂无活动记录"

    lines = [f"🏃 **Coros 活动列表** （共{total}条，第{page}页）\n"]
    for i, act in enumerate(data_list, 1):
        dist = format_distance(act.get("distance"))
        dur = format_duration(act.get("totalTime"))
        label_id = act.get("labelId", "")
        name = act.get("name", "未命名")
        st = sport_name(act.get("sportType", 0))
        date_str = datetime.fromtimestamp(act.get("startTime", 0)).strftime("%m/%d %H:%M")
        lines.append(f"**{i}.** {st} · {name}")
        lines.append(f"     {dist} · {dur} · {date_str}")
        lines.append(f"     ID: `{label_id}`")

    lines.append(f"\n发送 `/coros详情 <序号>` 查看详细信息")
    lines.append(f"发送 `/coros下载 <序号>` 下载活动文件")
    return "\n".join(lines)


def cmd_coros_detail(index_str):
    """获取活动详情"""
    global _coros_activity_cache
    try:
        idx = int(index_str) - 1
    except (ValueError, TypeError):
        return "❌ 请指定活动序号，如：/coros详情 1"

    if idx < 0 or idx >= len(_coros_activity_cache):
        return f"❌ 序号超出范围（当前缓存 {len(_coros_activity_cache)} 条），请先执行 /coros活动"

    act = _coros_activity_cache[idx]
    label_id = act.get("labelId", "")
    result = run_coros_cmd("detail", label_id)

    if not result["ok"]:
        return f"❌ 获取详情失败：{result['error']}"

    data = result.get("data", {})
    summary = data.get("summary", {})

    dist = format_distance(summary.get("distance", 0))
    dur = format_duration(summary.get("totalTime", 0))
    pace_str = format_duration(summary.get("avgPace", 0)) if summary.get("avgPace") else "—"
    hr_str = f"{summary.get('avgHr', '—')} bpm" if summary.get("avgHr") else "—"
    cal = f"{summary.get('calories', '—')} kcal" if summary.get("calories") else "—"
    elev = f"{summary.get('elevGain', '—')} m" if summary.get("elevGain") is not None else "—"
    cadence = f"{summary.get('avgCadence', '—')}" if summary.get("avgCadence") else "—"
    training_load = summary.get("trainingLoad", "—")

    lines = [
        f"📊 **{act.get('name', '活动详情')}**",
        f"",
        f"距离：{dist}",
        f"时长：{dur}",
        f"配速：{pace_str}",
        f"心率：{hr_str}",
        f"卡路里：{cal}",
        f"爬升：{elev}",
        f"步频：{cadence}",
        f"训练负荷：{training_load}",
        f"",
        f"发送 `/coros下载 {index_str}` 下载此活动的 .fit 文件",
    ]
    return "\n".join(lines)


def cmd_coros_download(index_str, file_type="fit"):
    """下载活动文件"""
    global _coros_activity_cache
    try:
        idx = int(index_str) - 1
    except (ValueError, TypeError):
        return "❌ 请指定活动序号，如：/coros下载 1"

    if idx < 0 or idx >= len(_coros_activity_cache):
        return f"❌ 序号超出范围，请先执行 /coros活动"

    if file_type not in ("fit", "tcx", "gpx", "csv"):
        file_type = "fit"

    act = _coros_activity_cache[idx]
    label_id = act.get("labelId", "")
    result = run_coros_cmd("download", label_id, "--type", file_type, "--output", str(WORK_DIR))

    if result["ok"]:
        filename = result.get("filename", "unknown")
        filepath = result.get("filePath", "")
        size_mb = ""
        if filepath and os.path.exists(filepath):
            size_mb = f" ({os.path.getsize(filepath) / (1024 * 1024):.1f} MB)"
        return f"✅ 已下载：`{filename}`{size_mb}\n存放于：{WORK_DIR}"
    return f"❌ 下载失败：{result['error']}"


def cmd_coros_help():
    return (
        "⌚ **Coros 指令**\n"
        "/coros活动 [页码] — 获取活动列表\n"
        "/coros详情 <序号> — 查看活动详细数据\n"
        "/coros下载 <序号> [fit|tcx|gpx|csv] — 下载活动文件\n"
        "/coros登录 <邮箱> <密码> — 配置 Coros 账号"
    )


# 指令路由
ROUTER = {
    "帮助": cmd_help,
    "计划": cmd_plan,
    "下载": cmd_download,
    "处理": cmd_process,
    "文件": cmd_files,
    "状态": cmd_status,
    "删除": cmd_delete,
    "coros帮助": cmd_coros_help,
    "coros活动": cmd_coros_activities,
    "coros详情": cmd_coros_detail,
    "coros下载": cmd_coros_download,
    "coros登录": cmd_coros_login,
}


def dispatch(text):
    """解析指令并执行"""
    text = text.strip()
    if text.startswith("/"):
        text = text[1:]

    parts = text.split(None, 1)
    cmd = parts[0] if parts else ""
    args_str = parts[1] if len(parts) > 1 else ""

    if cmd in ("下载", "download"):
        args = args_str.split()
        url = args[0] if args else ""
        filename = args[1] if len(args) > 1 else None
        return cmd_download(url, filename)

    if cmd in ("计划", "plan"):
        return cmd_plan(args_str)

    if cmd in ("日报",):
        return cmd_daily_report(args_str)

    if cmd in ("处理", "process"):
        args = args_str.split()
        filename = args[0] if args else ""
        op = args[1] if len(args) > 1 else "统计"
        return cmd_process(filename, op)

    if cmd in ("删除", "delete", "del", "rm"):
        return cmd_delete(args_str)

    if cmd in ("coros登录",):
        args = args_str.split()
        email = args[0] if args else ""
        password = args[1] if len(args) > 1 else ""
        return cmd_coros_login(email, password)

    if cmd in ("coros活动",):
        page = int(args_str) if args_str else 1
        return cmd_coros_activities(page)

    if cmd in ("coros详情",):
        return cmd_coros_detail(args_str)

    if cmd in ("coros下载",):
        args = args_str.split()
        index = args[0] if args else ""
        file_type = args[1] if len(args) > 1 else "fit"
        return cmd_coros_download(index, file_type)

    handler = ROUTER.get(cmd)
    if handler:
        return handler()

    return f"❓ 未知指令：`{cmd}`\n发送「帮助」查看可用指令"


# ── Flask 服务 ──────────────────────────────────────────────────

app = Flask(__name__)


@app.route("/feishu/event", methods=["POST"])
def handle_event():
    body = request.get_json(force=True)

    # URL 验证（飞书配置回调地址时会发送 challenge）
    if body.get("type") == "url_verification":
        token = body.get("token", "")
        challenge = body.get("challenge", "")
        # 简单验证：返回加密的 challenge
        if not is_encrypted(body):
            return jsonify({"challenge": challenge})
        # 如果开启了加密，需要解密后返回
        return jsonify({"challenge": challenge})

    # 事件回调
    header = body.get("header", {})
    event_type = header.get("event_type", "")

    if event_type == "im.message.receive_v1":
        event = body.get("event", {})
        message = event.get("message", {})
        msg_type = message.get("message_type", "")
        chat_type = message.get("chat_type", "")

        # 暂时未开启加密，直接取 content
        content_str = message.get("content", "{}")
        try:
            content = json.loads(content_str)
        except (json.JSONDecodeError, TypeError):
            return jsonify({"code": 0})

        text = content.get("text", "").strip()

        # 获取发送者 open_id 用于回复
        sender_id = event.get("sender", {}).get("sender_id", {}).get("open_id", "")

        if not text:
            return jsonify({"code": 0})

        print(f"[{datetime.now()}] 收到指令: {text} (from {sender_id})")

        result = dispatch(text)
        if sender_id:
            send_message(sender_id, result)

    return jsonify({"code": 0})


def is_encrypted(body):
    """检测是否开启了加密模式"""
    return "encrypt" in body


@app.route("/ping")
def ping():
    return "feishu-bot is running"


# ── 启动 ───────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f" 飞书 Bot 启动中...")
    print(f"   端口: {PORT}")
    print(f"   工作目录: {WORK_DIR}")
    app.run(host="0.0.0.0", port=PORT, debug=False)
